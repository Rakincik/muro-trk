import os
import argparse
import uuid
import datetime
import psycopg2
import openpyxl

def parse_name(full_name):
    if not full_name: return "", ""
    parts = str(full_name).strip().split()
    if len(parts) == 0: return "", ""
    if len(parts) == 1: return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def clean_phone(phone):
    if not phone: return None
    digits = "".join(c for c in str(phone) if c.isdigit())
    while digits.startswith("0"): digits = digits[1:]
    if len(digits) == 12 and digits.startswith("90"): digits = digits[2:]
    return digits if digits else None

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", required=True)
    parser.add_argument("--dbname", required=True)
    parser.add_argument("--user", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args()

    print("Connecting to HLL Database...")
    conn = psycopg2.connect(
        host=args.host, dbname=args.dbname, user=args.user, password=args.password
    )
    cur = conn.cursor()

    wb = openpyxl.load_workbook("dershaneonline.xlsx", data_only=True)

    # 1. IMPORT USERS
    print("\n--- IMPORTING USERS ---")
    s1 = wb["Sayfa1"]
    # Headers are at row 3, data starts at row 4 or 5.
    # Columns: M(12)=Ad_Soyad, N(13)=Telefon, O(14)=Eposta, P(15)=Rol
    for row in s1.iter_rows(min_row=4, values_only=True):
        if len(row) < 16: continue
        
        ad_soyad = row[12]
        telefon = row[13]
        eposta = row[14]
        rol = row[15]
        
        if not ad_soyad and not telefon: continue
        if ad_soyad == "Ad_Soyad": continue

        first_name, last_name = parse_name(ad_soyad)
        if not first_name: first_name = "Ogrenci"
        phone = clean_phone(telefon)
        email = str(eposta).strip() if eposta else f"{phone}@hll.muro.click" if phone else f"{uuid.uuid4().hex[:8]}@hll.muro.click"
        username = phone if phone else email.split('@')[0]
        password_hash = "AQAAAAEAACcQAAAAEFg8b7gB5v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6v5V6w==" # 123456
        
        role_str = "Admin"
        if str(rol).lower() == "öğrenci": role_str = "Student"
        elif str(rol).lower() == "eğitimci": role_str = "Teacher"

        cur.execute("SELECT \"Id\" FROM \"Users\" WHERE \"Phone\" = %s OR \"Email\" = %s", (phone, email))
        if not cur.fetchone():
            if args.execute:
                cur.execute(
                    "INSERT INTO \"Users\" (\"Id\", \"FirstName\", \"LastName\", \"Email\", \"Username\", \"Phone\", \"PasswordHash\", \"Role\", \"StudentType\", \"IsActive\", \"IsDeleted\", \"CreatedAt\") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (str(uuid.uuid4()), first_name, last_name, email, username, phone, password_hash, role_str, "Active", True, False, datetime.datetime.now())
                )
            print(f"Added User: {first_name} {last_name} ({phone}) - Role: {role_str}")

    # 2. IMPORT COURSES
    print("\n--- IMPORTING COURSES ---")
    if "Sayfa2" in wb.sheetnames:
        s2 = wb["Sayfa2"]
        for row in s2.iter_rows(min_row=2, values_only=True):
            ders = row[0]
            if not ders or str(ders).strip() == "": continue

            course_name = str(ders).strip()
            cur.execute("SELECT \"Id\" FROM \"Courses\" WHERE \"Title\" = %s", (course_name,))
            if not cur.fetchone():
                if args.execute:
                    cur.execute(
                        "INSERT INTO \"Courses\" (\"Id\", \"Title\", \"IsDeleted\", \"CreatedAt\") VALUES (%s, %s, %s, %s)",
                        (str(uuid.uuid4()), course_name, 0, False, datetime.datetime.now())
                    )
                print(f"Added Course: {course_name}")

    if args.execute:
        conn.commit()
        print("\n[+] SUCCESS: Data imported successfully!")
    else:
        print("\n[!] DRY RUN: Run with --execute to commit changes to DB")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
